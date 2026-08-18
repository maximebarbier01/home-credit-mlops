from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.db.database import init_db
from app.db.repository import save_api_call_log, save_prediction_log
from home_credit_mlops.monitoring.production import (
    load_api_call_logs,
    load_prediction_logs,
    load_production_inputs,
    load_production_outputs,
    production_feature_frame,
)
from home_credit_mlops.monitoring.report import build_monitoring_report


def _seed_monitoring_database(database_url: str) -> None:
    init_db(database_url)
    payload = {
        "AGE_YEARS": 35.0,
        "AMT_INCOME_TOTAL": 50_000.0,
        "AMT_CREDIT": 200_000.0,
        "CODE_GENDER": "F",
    }
    response = {
        "default_probability": 0.42,
        "business_threshold": 0.22,
        "predicted_default": 1,
        "credit_decision": "refused",
    }
    save_prediction_log(payload, response, latency_ms=12.5)
    save_api_call_log(
        method="POST",
        path="/predict",
        status_code=200,
        latency_ms=12.5,
        request_payload=payload,
    )
    save_api_call_log(
        method="POST",
        path="/predict",
        status_code=422,
        latency_ms=4.0,
        request_payload={"AGE_YEARS": -5},
        error_type="http_422",
        error_message="Unprocessable Entity",
    )


def test_load_logs_and_flatten_production_features(tmp_path: Path) -> None:
    database_url = f"sqlite:///{(tmp_path / 'monitoring.db').as_posix()}"
    _seed_monitoring_database(database_url)

    api_calls = load_api_call_logs(database_url)
    prediction_logs = load_prediction_logs(database_url)
    production_features = production_feature_frame(prediction_logs)
    production_inputs = load_production_inputs(database_url)
    production_outputs = load_production_outputs(database_url)

    assert len(api_calls) == 2
    assert len(prediction_logs) == 1
    assert production_features.loc[0, "AGE_YEARS"] == 35.0
    assert production_inputs.loc[0, "AGE_YEARS"] == 35.0
    assert production_outputs.loc[0, "predicted_default"] == 1


def test_build_monitoring_report_exports_workbook_html_and_plots(tmp_path: Path) -> None:
    database_url = f"sqlite:///{(tmp_path / 'monitoring.db').as_posix()}"
    _seed_monitoring_database(database_url)
    reference_path = tmp_path / "reference.parquet"
    pd.DataFrame(
        {
            "SK_ID_CURR": [1, 2, 3],
            "TARGET": [0, 1, 0],
            "AGE_YEARS": [30.0, 40.0, 50.0],
            "AMT_INCOME_TOTAL": [40_000.0, 60_000.0, 70_000.0],
            "AMT_CREDIT": [180_000.0, 220_000.0, 250_000.0],
            "CODE_GENDER": ["F", "M", "F"],
        }
    ).to_parquet(reference_path, index=False)

    report = build_monitoring_report(
        database_url=database_url,
        reference_data_path=reference_path,
        output_dir=tmp_path / "report",
        target_column="TARGET",
        id_column="SK_ID_CURR",
        top_drift_features=5,
        latency_warning_ms=10.0,
        error_rate_warning=0.10,
    )

    assert report.workbook_path.exists()
    assert report.html_path.exists()
    assert {plot.name for plot in report.plots} >= {
        "score_distribution.png",
        "decision_distribution.png",
        "latency_distribution.png",
        "top_drift_features.png",
    }

    workbook = load_workbook(report.workbook_path)
    assert {"api_summary", "prediction_summary", "drift_summary"}.issubset(workbook.sheetnames)
    assert {"production_inputs_sample", "production_outputs_sample"}.issubset(workbook.sheetnames)
