from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from app.db.database import init_db
from app.db.repository import save_api_call_log, save_prediction_log
from home_credit_mlops.performance.report import (
    build_performance_report,
    identify_bottlenecks,
    summarize_latency,
)


def _seed_performance_database(database_url: str) -> None:
    init_db(database_url)
    payload = {
        "AGE_YEARS": 35.0,
        "AMT_INCOME_TOTAL": 50_000.0,
    }
    response = {
        "default_probability": 0.42,
        "business_threshold": 0.22,
        "predicted_default": 1,
        "credit_decision": "refused",
    }
    save_prediction_log(payload, response, latency_ms=18.0)
    save_api_call_log(method="POST", path="/predict", status_code=200, latency_ms=30.0)
    save_api_call_log(
        method="POST",
        path="/predict",
        status_code=422,
        latency_ms=5.0,
        request_payload={"AGE_YEARS": -5},
        error_type="http_422",
    )


def test_summarize_latency_handles_empty_frame() -> None:
    summary = summarize_latency(frame=pd.DataFrame(), source="api")

    assert summary.loc[0, "source"] == "api"
    assert summary.loc[0, "count"] == 0
    assert summary.loc[0, "latency_p95_ms"] == 0.0


def test_build_performance_report_exports_workbook_and_markdown(tmp_path: Path) -> None:
    database_url = f"sqlite:///{(tmp_path / 'performance.db').as_posix()}"
    _seed_performance_database(database_url)

    report = build_performance_report(
        database_url=database_url,
        output_dir=tmp_path / "performance_report",
        latency_warning_ms=20.0,
        error_rate_warning=0.10,
    )

    assert report.workbook_path.exists()
    assert report.markdown_path.exists()
    assert "Rapport d'analyse" in report.markdown_path.read_text(encoding="utf-8")

    workbook = load_workbook(report.workbook_path)
    assert {
        "api_summary",
        "model_latency_summary",
        "bottlenecks",
        "optimization_decisions",
    }.issubset(workbook.sheetnames)


def test_identify_bottlenecks_flags_latency_and_error_rate() -> None:
    bottlenecks = identify_bottlenecks(
        api_summary=pd.DataFrame(
            [
                {
                    "latency_p95_ms": 1200.0,
                    "error_rate": 0.20,
                }
            ]
        ),
        model_latency_summary=pd.DataFrame([{"latency_p95_ms": 100.0}]),
        latency_warning_ms=1000.0,
        error_rate_warning=0.05,
    )

    assert "Latence API p95 élevée" in bottlenecks["bottleneck"].tolist()
    assert "Taux d'erreur HTTP" in bottlenecks["bottleneck"].tolist()
